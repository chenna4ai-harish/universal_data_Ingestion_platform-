"""
Run this script once to generate the XLSX test files:
  python test_data/generate_xlsx.py
Requires: openpyxl  (pip install openpyxl)
"""
import os
import openpyxl
from openpyxl import Workbook

OUT = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# 1. customers_exact.xlsx  — exact column names, multi-sheet (only Sheet1 used)
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Customers"
ws.append([
    "account_number", "company_name", "duns_number", "government_id",
    "country_code", "address_1", "city", "state", "postcode"
])
rows = [
    ("CUST-050","Upsilon Corp","901234567","US11223344","US","99 Broadway","New York","NY","10005"),
    ("CUST-051","Phi Logistics","012345678","GB55443322","GB","7 Churchill Way","Cardiff","Wales","CF10 2HH"),
    ("CUST-052","Chi Services Pty","123456780","AU66778899","AU","1 Pacific Hwy","Sydney","NSW","2060"),
    ("CUST-053","Psi Imports","234567891","DE44332211","DE","Schillerstr. 3","Stuttgart","BW","70173"),
    ("CUST-054","Omega Trade LLC","345678902","US99887711","US","600 Market St","San Francisco","CA","94105"),
]
for r in rows:
    ws.append(r)
wb.save(os.path.join(OUT, "customers_exact.xlsx"))
print("Created customers_exact.xlsx")

# ---------------------------------------------------------------------------
# 2. invoices_exact.xlsx  — exact column names
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Invoices"
ws.append([
    "account_number", "invoice_number", "invoice_date", "due_date",
    "invoice_amount", "paid_date", "paid_amount", "payment_terms",
    "invoice_type", "currency", "doe"
])
rows = [
    ("CUST-050","INV-2024-0050","2024-01-08","2024-02-07",32000.00,"2024-02-05",32000.00,"Net 30","PAID","USD","2024-03-31"),
    ("CUST-051","INV-2024-0051","2024-01-15","2024-04-14",18500.00,"","",            "Net 90","OPEN","GBP","2024-03-31"),
    ("CUST-052","INV-2024-0052","2024-02-01","2024-03-02",5600.00, "2024-02-28",5600.00,"Net 30","PAID","AUD","2024-03-31"),
    ("CUST-053","INV-2024-0053","2024-02-10","2024-03-11",11200.00,"","",            "Net 30","OPEN","EUR","2024-03-31"),
    ("CUST-054","INV-2024-0054","2024-03-01","2024-04-30",64000.00,"","",            "Net 60","OPEN","USD","2024-03-31"),
    ("CUST-050","INV-2024-0055","2024-03-15","2024-04-14",9100.00, "2024-04-12",9100.00,"Net 30","PAID","USD","2024-03-31"),
]
for r in rows:
    ws.append(r)
wb.save(os.path.join(OUT, "invoices_exact.xlsx"))
print("Created invoices_exact.xlsx")

# ---------------------------------------------------------------------------
# 3. invoices_fuzzy_headers.xlsx  — fuzzy column names to test mapper
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "AR Data"
ws.append([
    "Client No", "Bill Number", "Billing Date", "Due By Date",
    "Gross Amt", "Settlement Date", "Receipt Amount", "Trade Terms",
    "Doc Type", "ISO Currency", "Cutoff Date"
])
rows = [
    ("CUST-050","INV-2024-0060","2024-01-20","2024-02-19",14500.00,"2024-02-17",14500.00,"Net 30","PAID","USD","2024-03-31"),
    ("CUST-051","INV-2024-0061","2024-02-05","2024-05-05",72000.00,"","",            "Net 90","OPEN","GBP","2024-03-31"),
    ("CUST-052","INV-2024-0062","2024-02-12","2024-03-13",4200.00, "2024-03-11",4200.00,"Net 30","PAID","AUD","2024-03-31"),
    ("CUST-053","INV-2024-0063","2024-03-02","2024-04-01",8900.00, "","",           "Net 30","OPEN","EUR","2024-03-31"),
]
for r in rows:
    ws.append(r)
wb.save(os.path.join(OUT, "invoices_fuzzy_headers.xlsx"))
print("Created invoices_fuzzy_headers.xlsx")

print("\nAll XLSX files generated in:", OUT)
